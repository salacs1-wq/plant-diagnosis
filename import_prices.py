import openpyxl
from db import get_connection


XLSX_PATH = "prices_source.xlsx"
SUPPLIER_NAME = "current"
CURRENCY = "HUF"


def to_float(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(" ", "").replace("\u00a0", "")
    text = text.replace(".", "")
    text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def import_prices():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    conn = get_connection()
    cur = conn.cursor()

    # Fejléc beolvasás
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    header_map = {h: i for i, h in enumerate(headers)}

    name_idx = header_map.get("Név")
    pack_idx = header_map.get("Fk")
    price_idx = header_map.get("Kedvezm.ár")

    if name_idx is None or price_idx is None:
        raise ValueError("Hiányzik a szükséges oszlop: 'Név' vagy 'Kedvezm.ár'.")

    for row in ws.iter_rows(min_row=2, values_only=True):
        product_name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
        pack_size = str(row[pack_idx]).strip() if pack_idx is not None and row[pack_idx] is not None else None
        price_net = to_float(row[price_idx])

        if not product_name or price_net is None:
            continue

        # products.termek alapján párosítunk
        cur.execute(
            """
            SELECT id, termek
            FROM products
            WHERE lower(trim(termek)) = lower(trim(?))
            LIMIT 1
            """,
            (product_name,)
        )
        product_row = cur.fetchone()

        if not product_row:
            continue

        product_id = product_row["id"]
        price_calc = round(price_net * 1.03, 2)

        cur.execute(
            """
            INSERT INTO product_prices (
                product_id,
                supplier,
                pack_size,
                price_net,
                price_calc,
                currency,
                valid_from,
                valid_to,
                is_current
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                product_id,
                SUPPLIER_NAME,
                pack_size,
                price_net,
                price_calc,
                CURRENCY,
                None,
                None,
                1
            )
        )

    conn.commit()
    conn.close()


if __name__ == "__main__":
    import_prices()
    print("prices import done")
